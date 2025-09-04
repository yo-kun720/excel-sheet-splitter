### --- ここから実装 ---


import pandas as pd
import zipfile
import io
import re
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from copy import copy
from openpyxl.cell.cell import Cell


def sanitize_sheet_name(sheet_name: str) -> str:
    """
    OSで禁止されている文字を除去し、255文字以内にする。
    """
    # Windows禁止文字: \\/:*?"<>|
    sanitized = re.sub(r'[\\/:*?"<>|]', '_', sheet_name)
    return sanitized[:255]


def _is_formula_cell(cell: Cell) -> bool:
    try:
        if cell is None:
            return False
        if cell.data_type == 'f':
            return True
        if isinstance(cell.value, str) and cell.value.startswith('='):
            return True
        return False
    except Exception:
        return False


def split_excel_to_zip(file) -> bytes:
    """
    Excelファイルを受け取り、各シートを個別のExcelファイルに分割し、
    それらをZipバイト列として返す。元の書式設定を保持する。
    - 数式セルは『計算済みの表示値』に固定（キャッシュが無ければ数式文字列をプレーンテキスト化）
    - 削除/非表示/不正なシートは除外
    Args:
        file: BytesIOまたはアップロードファイルオブジェクト
    Returns:
        bytes: Zipファイルのバイト列
    Raises:
        ValueError: Excelファイルが不正な場合
    """
    try:
        # ファイルをバイトに確定させ、2回読み込めるようにする
        if hasattr(file, 'read'):
            bytes_data = file.read()
        else:
            bytes_data = file if isinstance(file, (bytes, bytearray)) else bytes(file)

        # 値読み込み（data_only=True）→ 数式セルは計算結果（キャッシュ値）を取得
        workbook_values = load_workbook(io.BytesIO(bytes_data), data_only=True)
        # 数式読み込み（data_only=False）→ 数式の有無判定に使用
        workbook_formulas = load_workbook(io.BytesIO(bytes_data), data_only=False)
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みに失敗しました: {e}")

    # 実在し処理対象のワークシート収集
    from openpyxl.worksheet.worksheet import Worksheet
    valid_worksheets = []
    excluded_sheets = []

    # タイトル→Worksheet（数式側）対応表
    title_to_ws_formula = {ws.title: ws for ws in workbook_formulas.worksheets}

    for ws in workbook_values.worksheets:
        try:
            if not isinstance(ws, Worksheet):
                excluded_sheets.append(f"{getattr(ws, 'title', 'Unknown')} (Worksheet型ではない)")
                continue
            if not hasattr(ws, 'title') or not ws.title:
                excluded_sheets.append("(無題/タイトル欠落)")
                continue

            state_value = getattr(ws, 'sheet_state', getattr(ws, 'state', None))
            if state_value in ("veryHidden", "hidden"):
                excluded_sheets.append(f"{ws.title} (非表示シート: {state_value})")
                continue

            if not hasattr(ws, 'max_row') or not hasattr(ws, 'max_column'):
                excluded_sheets.append(f"{ws.title} (寸法属性欠落)")
                continue

            _ = ws.cell(row=1, column=1)

            valid_worksheets.append(ws)
        except Exception as e:
            excluded_sheets.append(f"{getattr(ws, 'title', 'Unknown')} (アクセス失敗: {e})")
            continue

    if excluded_sheets:
        print(f"除外されたシート: {', '.join(excluded_sheets)}")
    print(f"対象シート: {[ws.title for ws in valid_worksheets]}")

    if not valid_worksheets:
        raise ValueError("有効なシートが見つかりませんでした。")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for ws_values in valid_worksheets:
            sheet_name = ws_values.title
            try:
                # 各シート処理ごとにワークブックを再読込（体裁を完全保持するため）
                wb_vals = load_workbook(io.BytesIO(bytes_data), data_only=True)
                wb_form = load_workbook(io.BytesIO(bytes_data), data_only=False)

                if sheet_name not in wb_form.sheetnames or sheet_name not in wb_vals.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")

                ws_v = wb_vals[sheet_name]
                ws_f = wb_form[sheet_name]

                # 数式をその場で値に置換（対象シートのみ）
                for row in ws_f.iter_rows():
                    for cell_f in row:
                        cell_v = ws_v.cell(row=cell_f.row, column=cell_f.column)
                        if _is_formula_cell(cell_f):
                            evaluated = cell_v.value
                            if evaluated is None:
                                formula_text = cell_f.value if isinstance(cell_f.value, str) else ""
                                cell_f.value = "'" + str(formula_text)
                            else:
                                cell_f.value = evaluated

                # 対象シート以外を削除
                for other in list(wb_form.sheetnames):
                    if other != sheet_name:
                        ws_del = wb_form[other]
                        # veryHidden/hidden等も含めて削除
                        wb_form.remove(ws_del)

                # 保存
                sheet_buffer = io.BytesIO()
                wb_form.save(sheet_buffer)
                sheet_buffer.seek(0)

                sheet_filename = sanitize_sheet_name(sheet_name) + ".xlsx"
                zipf.writestr(sheet_filename, sheet_buffer.getvalue())

            except Exception as e:
                raise ValueError(f"シート '{sheet_name}' の処理中にエラー: {e}")

    zip_buffer.seek(0)
    return zip_buffer.getvalue()

if __name__ == "__main__":
    # テスト用: コマンドラインから実行した場合の動作例
    with open("test.xlsx", "rb") as f:
        zip_bytes = split_excel_to_zip(f)
    with open("sheets.zip", "wb") as f:
        f.write(zip_bytes) 